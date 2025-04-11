import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def calculate_commission(row):
    # Проверяем, что это операция "Оплата"
    if row['Тип операции'] != 'Оплата':
        return 0  # Возвращаем 0 для отмен или других типов операций

    # Поиск колонки с типом карты
    card_type_column = None
    possible_names = [' ПС', 'PS', 'Тип карты', 'Card Type', 'Payment System']

    for name in possible_names:
        if name in row.index:
            card_type_column = name
            break

    if not card_type_column:
        raise KeyError("Не найдена колонка с типом карты")

    card_type = row[card_type_column]
    amount = row['Сумма']

    commission_rates = {
        'MASTER_CARD': 0.0185,
        'VISA': 0.0133,
        'CHINA_UNION_PAY': 0.0188,
        'WORLD': 0.0118  # Карта мир
    }

    rate = commission_rates.get(card_type, 0.01)
    return round(amount * rate, 2)


file_path = 'Отчет за 12.03.xlsx'
sheet_name = 'Лист2'

try:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print("Колонки в файле:", df.columns.tolist())

    # Добавляем колонки
    df['Комиссия (расчет)'] = df.apply(calculate_commission, axis=1)
    df['Разница (F - U)'] = df.apply(
        lambda row: (row['Комиссия'] - row['Комиссия (расчет)']) if row['Тип операции'] == 'Оплата' else 0,
        axis=1
    )
    df['Разница (F - U)'] = df['Разница (F - U)'].round(2)

    # Сохранение
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Окрашивание
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Зеленый для расчетной комиссии
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    for row in ws.iter_rows(min_row=2, min_col=21, max_col=21):
        for cell in row:
            if cell.value is not None and cell.value != 0:
                cell.fill = green_fill

    # Цвета для разницы
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green_diff_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')  # Серый для отмен

    for row in ws.iter_rows(min_row=2, min_col=22, max_col=22):
        for cell in row:
            if cell.value is not None:
                if cell.value == 0:
                    cell.fill = gray_fill
                elif cell.value < 0:
                    cell.fill = red_fill
                else:
                    cell.fill = green_diff_fill

    wb.save(file_path)
    print("Обработка завершена. Добавлены колонки с учетом типа операции.")

except Exception as e:
    print(f"Ошибка: {e}")
    print("Проверьте названия колонок. Нужны: 'Тип операции', ' ПС', 'Сумма', 'Комиссия'")